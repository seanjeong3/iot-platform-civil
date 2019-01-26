import xlrd
import xml.etree.ElementTree as ET
import xml.dom.minidom as minidom
import openpyxl
from brim_to_csi_dictionary import brim_to_csi_dictionary as btocDict
from csi_to_brim_dictionary import csi_to_brim_dictionary as ctobDict

templateFileName = 'brim_lib/trb_template.xlsx'
inputFileName = 'brim_lib/trb_brim_fem_retrieved.xml'
outputFileName = 'brim_lib/trb_brim_excel.xlsx'

#############################################

def GetAttribute(items):
	attribute = {}
	for item in items:
		attribute[item[0]] = item[1]
	return attribute

def GetParameter(params):
	parameter = {}
	paramAttr = {}
	if params :
		for param in params:
			paramAttr = GetAttribute(param.items())
			paramVal = paramAttr['V']
			if 'Type' not in paramAttr:
				paramVal = float(paramVal)
			parameter[paramAttr['N']] = paramVal
	return parameter

def ScanAttributes(attributeDict, sheet):
	columnDict = {}
	ncol = sheet.max_column
	for i in range(0, ncol):
		att = sheet.cell(row = 2, column = i+1).value
		if att in attributeDict:
			if attributeDict[att] == 'N':
				columnDict[attributeDict[att]] = i+1
			else:
				columnDict[attributeDict[att].lower()] = i+1
	return columnDict 

def ParseData(data, book):
	if data['T'] in ctobDict:
		for sheetName in ctobDict[data['T']]:

			if sheetName == 'Area Overwrites - Joint Offsets' and 'offset1' in data:
				sheet = book.get_sheet_by_name(sheetName)
				nrow = sheet.max_row
				i = 1
				while i <= data['numjoints']:
					if i % 4 == 1:
						nrow = nrow + 1
						sheet.cell(row = nrow, column = 1).value = data['N']
					if i == 1:
						sheet.cell(row = nrow, column = 2).value = 'Object'
					att = 'offset' + str(i)
					sheet.cell(row = nrow,  column = 3+(i-1)%4).value = data[att]
					i = i + 1
			elif sheetName == 'Connectivity - Area' and 'node1' in data:
				sheet = book.get_sheet_by_name(sheetName)
				nrow = sheet.max_row
				i = 1
				while i <= data['numjoints']:
					if i % 4 == 1:
						nrow = nrow + 1
						sheet.cell(row = nrow, column = 1).value = data['N']
					if i == 1:
						sheet.cell(row = nrow, column = 2).value = data['numjoints']
					att = 'node' + str(i)
					sheet.cell(row = nrow,  column = 3+(i-1)%4).value = data[att]
					i = i + 1
			elif sheetName == 'Bridge Layout Line 2 - Horiz' and 'azimuth' in data:
				intersec = list(set(data.keys()) & set(btocDict[data['T']][sheetName].keys()))
				if len(intersec) > 1:
					sheet = book.get_sheet_by_name(sheetName)
					columnDict = ScanAttributes(ctobDict[data['T']][sheetName], sheet)
					nrow = sheet.max_row
					for d in intersec:
						if d in columnDict:
							if d == 'azimuth':
								sheet.cell(row = nrow + 1, column = columnDict[d]).value = 'N' + str(int(data[d])*10000) + 'E'
							else:
								sheet.cell(row = nrow + 1, column = columnDict[d]).value = data[d]

			else:
				if sheetName == 'Function - PSD - User' and data['N'] != 'UNIFPSD': continue
				if sheetName == 'Function - Resp Spect - User' and data['N'] != 'UNIFRS': continue
				if sheetName == 'Function - Steady State - User' and data['N'] != 'UNIFSS': continue
				if sheetName == 'Function - Time History - User' and data['N'] not in ['RAMPTH', 'UNIFTH']: continue
				if sheetName == 'MatProp 03a - Steel Data' and data['type'] != 'Steel': continue
				if sheetName == 'MatProp 03b - Concrete Data' and data['type'] != 'Concrete': continue
				if sheetName == 'MatProp 03e - Rebar Data' and data['type'] != 'Rebar': continue
				if sheetName == 'MatProp 03f - Tendon Data' and data['type'] != 'Tendon': continue
				intersec = list(set(data.keys()) & set(btocDict[data['T']][sheetName].keys()))
				if len(intersec) > 1:
					sheet = book.get_sheet_by_name(sheetName)
					columnDict = ScanAttributes(ctobDict[data['T']][sheetName], sheet)
					nrow = sheet.max_row
					for d in intersec:
						if d in columnDict:
							sheet.cell(row = nrow + 1, column = columnDict[d]).value = data[d]


def ParseObject(corrObj, book):
	attribute = {}
	childObjs = {}
	data = {}
	attribute = GetAttribute(corrObj.items())
	parameter = GetParameter(corrObj.findall('P'))

	data = attribute.copy()
	data.update(parameter)

	ParseData(data, book)

	for childObj in corrObj.findall('O'):
		ParseObject(childObj, book)
		



#############################################

if __name__ == '__main__':


	root = ET.parse(inputFileName).getroot()
	book = openpyxl.load_workbook(templateFileName)
	ParseObject(root, book)
	book.save(outputFileName)
	print outputFileName
